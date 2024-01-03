from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Pega a nova planilha e transforma em um workbook
wb = load_workbook("data/ValorVendaFabricantePorAno.xlsx")
sheet = wb["Relatorio"]

# Configura coluna e linha maxima e minima
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Cria o grafico
barchart = BarChart()

# Cria primeira metrica do grafico
data = Reference(
    sheet,
    min_col = min_column + 1,
    max_col = max_column,
    min_row = min_row,
    max_row = max_row
)

# Cria primeira metrica do grafico
categories = Reference(
    sheet,
    min_col = min_column,
    max_col = min_column,
    min_row = min_row + 1,
    max_row = max_row
)

# Adiciona as metricas no grafico
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# Adiciona Grafico na Planilha
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2

# Salva o Grafico em uma nova planilha
wb.save("data/Grafico.xlsx")